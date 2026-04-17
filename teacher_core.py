# teacher_core.py  –  logic dùng chung, KHÔNG cần Anthropic API
import re, io, unicodedata, difflib
from datetime import datetime, timedelta, date as date_type
from collections import defaultdict
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Bảng môn học theo từng cấp ───────────────────────────────────────────────

# Tiểu học (khối 1–5)
SUBJECT_MAP_TH = {
    "tiếng việt":"TIENGVIET","tieng viet":"TIENGVIET","tv":"TIENGVIET",
    "t.v":"TIENGVIET","tviet":"TIENGVIET","tviệt":"TIENGVIET","t.việt":"TIENGVIET",
    "toán":"TOAN","toán học":"TOAN","toan":"TOAN",
    "tiếng anh":"ANH","anh":"ANH","ngoại ngữ":"ANH","ngoai ngu":"ANH","english":"ANH",
    "tự nhiên và xã hội":"TUNHIENVAXAHOI","tu nhien va xa hoi":"TUNHIENVAXAHOI",
    "tnxh":"TUNHIENVAXAHOI","xã hội":"TUNHIENVAXAHOI","xa hoi":"TUNHIENVAXAHOI",
    "tự nhiên":"TUNHIENVAXAHOI","tu nhien":"TUNHIENVAXAHOI",
    "đạo đức":"DAODUC","dao duc":"DAODUC","dd":"DAODUC","đ.đức":"DAODUC","d.duc":"DAODUC",
    "tin học và công nghệ":"THCN","tin hoc va cong nghe":"THCN",
    "tin":"THCN","công nghệ":"THCN","cong nghe":"THCN","thcn":"THCN",
    "khoa học":"KHOAHOC","khoa hoc":"KHOAHOC","khtn":"KHOAHOC","khoa":"KHOAHOC",
    "giáo dục thể chất":"GDTC","giao duc the chat":"GDTC","thể dục":"GDTC","the duc":"GDTC","gdtc":"GDTC",
    "lịch sử và địa lí":"LICHSUDIALI","lich su va dia li":"LICHSUDIALI",
    "lịch sử và địa lý":"LICHSUDIALI","lịch sử":"LICHSUDIALI(SU)","lich su":"LICHSUDIALI(SU)",
    "địa lý":"LICHSUDIALI(DIA)","địa lí":"LICHSUDIALI(DIA)","dia ly":"LICHSUDIALI(DIA)","dia li":"LICHSUDIALI(DIA)",
    "sử":"LICHSUDIALI(SU)","su":"LICHSUDIALI(SU)","địa":"LICHSUDIALI(DIA)","dia":"LICHSUDIALI(DIA)",
    "lichsudialy":"LICHSUDIALI","ls&đl":"LICHSUDIALI","ls & đl":"LICHSUDIALI","lsdl":"LICHSUDIALI",
    "hoạt động ngoài giờ lên lớp":"HDNGLL","hoat dong ngoai gio len lop":"HDNGLL",
    "hdngll":"HDNGLL","ngoài giờ":"HDNGLL","ngoai gio":"HDNGLL",
    "tiếng dân tộc thiểu số":"TDTTS","tieng dan toc thieu so":"TDTTS","tdtts":"TDTTS",
    "nghệ thuật":"NGHETHUAT","nghe thuat":"NGHETHUAT",
    "âm nhạc":"NGHETHUAT(NHAC)","am nhac":"NGHETHUAT(NHAC)","nhạc":"NGHETHUAT(NHAC)","nhac":"NGHETHUAT(NHAC)",
    "mĩ thuật":"NGHETHUAT(MT)","mỹ thuật":"NGHETHUAT(MT)","my thuat":"NGHETHUAT(MT)","mi thuat":"NGHETHUAT(MT)",
    "họa":"NGHETHUAT(MT)","hoa":"NGHETHUAT(MT)",
    "hoạt động trải nghiệm":"HDTN","hoat dong trai nghiem":"HDTN","hdtn":"HDTN",
    "hđtn":"HDTN","trải nghiệm":"HDTN",
}

# Môn chính khóa GVCN Tiểu học theo khối
# GVCN dạy toàn bộ các môn này cho lớp mình chủ nhiệm
_TH_GVCN_SUBJECTS = {
    1: ["TOAN","TIENGVIET","TUNHIENVAXAHOI","DAODUC","HDTN"],
    2: ["TOAN","TIENGVIET","TUNHIENVAXAHOI","DAODUC","HDTN"],
    3: ["TOAN","TIENGVIET","TUNHIENVAXAHOI","DAODUC","HDTN"],
    4: ["TOAN","TIENGVIET","KHOAHOC","LICHSUDIALI","DAODUC","HDTN"],
    5: ["TOAN","TIENGVIET","KHOAHOC","LICHSUDIALI","DAODUC","HDTN"],
}


# THCS (khối 6–9)
SUBJECT_MAP_THCS = {
    "ngữ văn":"NGUVAN","ngữ văn học":"NGUVAN","van":"NGUVAN","nguvan":"NGUVAN","nv":"NGUVAN",
    "ngu van":"NGUVAN",
    "toán":"TOAN","toán học":"TOAN","toan":"TOAN",
    "nội dung giáo dục của địa phương":"NDGDDP",
    "nội dung giáo dục địa phương":"NDGDDP","giáo dục địa phương":"NDGDDP",
    "giao duc dia phuong":"NDGDDP","gdđp":"NDGDDP","gddp":"NDGDDP",
    "gd dp":"NDGDDP","gd11dp":"NDGDDP","nd gd dp":"NDGDDP",
    "tiếng anh":"ANH","anh":"ANH","ngoại ngữ":"ANH","ngoai ngu":"ANH","english":"ANH",
    "nn1":"ANH","nn2":"ANH",
    "công nghệ":"CONGNGHE","cong nghe":"CONGNGHE","c.nghệ":"CONGNGHE","c.nghe":"CONGNGHE",
    "cn":"CONGNGHE",
    "tin học":"TINHOC","tin hoc":"TINHOC","tin":"TINHOC","tinhoc":"TINHOC",
    "giáo dục công dân":"GDCD","giao duc cong dan":"GDCD","gdcd":"GDCD",
    "giáo dục thể chất":"GDTC","giao duc the chat":"GDTC","thể dục":"GDTC",
    "the duc":"GDTC","gdtc":"GDTC","td":"GDTC",
    "hoạt động ngoài giờ lên lớp":"HDNGLL","hoat dong ngoai gio len lop":"HDNGLL",
    "hdngll":"HDNGLL","ngoài giờ":"HDNGLL","ngoai gio":"HDNGLL",
    "tiếng dân tộc thiểu số":"TDTTS","tieng dan toc thieu so":"TDTTS","tdtts":"TDTTS",
    "nghề phổ thông":"NGHEPHOTHONG","nghe pho thong":"NGHEPHOTHONG","nghề":"NGHEPHOTHONG",
    "nghệ thuật (âm nhạc)":"NGHETHUAT(NHAC)","am nhac":"NGHETHUAT(NHAC)","âm nhạc":"NGHETHUAT(NHAC)",
    "nhạc":"NGHETHUAT(NHAC)","nhac":"NGHETHUAT(NHAC)","nghệ thuật (mĩ thuật)":"NGHETHUAT(MT)",
    "mỹ thuật":"NGHETHUAT(MT)","mĩ thuật":"NGHETHUAT(MT)","my thuat":"NGHETHUAT(MT)","mi thuat":"NGHETHUAT(MT)",
    "nghệ thuật":"NGHETHUAT","nghe thuat":"NGHETHUAT",
    "hoạt động trải nghiệm":"TNHN","hoat dong trai nghiem":"TNHN","hdtn":"TNHN",
    "hđtn":"TNHN","tnhn":"TNHN","hđ trải nghiệm":"TNHN",
    "tiếng pháp":"TIENGPHAP","tieng phap":"TIENGPHAP","pháp":"TIENGPHAP",
    "tiếng nga":"TIENGNGA","tieng nga":"TIENGNGA",
    "tiếng nhật":"TIENGNHAT","tieng nhat":"TIENGNHAT",
    "tiếng trung":"TIENGTRUNG","tieng trung":"TIENGTRUNG",
    "tiếng hàn":"TIENGHAN","tieng han":"TIENGHAN",
    # Khoa học tự nhiên — phân môn
    "khoa học tự nhiên (lí)":"KHTN(VATLY)","vật lý":"KHTN(VATLY)","vat ly":"KHTN(VATLY)",
    "vật lí":"KHTN(VATLY)","vat li":"KHTN(VATLY)","lý":"KHTN(VATLY)","lí":"KHTN(VATLY)",
    "ly":"KHTN(VATLY)","li":"KHTN(VATLY)","vatly":"KHTN(VATLY)","vl":"KHTN(VATLY)",
    "khoa học tự nhiên (hóa)":"KHTN(HOAHOC)","hóa học":"KHTN(HOAHOC)","hoa hoc":"KHTN(HOAHOC)",
    "hóa":"KHTN(HOAHOC)","hoá":"KHTN(HOAHOC)","hoa":"KHTN(HOAHOC)","hoahoc":"KHTN(HOAHOC)",
    "hh":"KHTN(HOAHOC)","hoá học":"KHTN(HOAHOC)",
    "khoa học tự nhiên (sinh)":"KHTN(SINH)","sinh học":"KHTN(SINH)","sinh hoc":"KHTN(SINH)",
    "sinh":"KHTN(SINH)",
    "khoa học tự nhiên":"KHTN","khoa hoc tu nhien":"KHTN","khtn":"KHTN",
    # Lịch sử & Địa lí — phân môn
    "lịch sử và địa lí (địa)":"LICHSUDIALI(DIA)","lịch sử và địa lý (địa)":"LICHSUDIALI(DIA)",
    "địa lí":"LICHSUDIALI(DIA)","địa lý":"LICHSUDIALI(DIA)",
    "dia li":"LICHSUDIALI(DIA)","dia ly":"LICHSUDIALI(DIA)","địa":"LICHSUDIALI(DIA)","dia":"LICHSUDIALI(DIA)",
    "lịch sử và địa lí (sử)":"LICHSUDIALI(SU)","lịch sử và địa lý (sử)":"LICHSUDIALI(SU)",
    "lịch sử":"LICHSUDIALI(SU)","lich su":"LICHSUDIALI(SU)","sử":"LICHSUDIALI(SU)","su":"LICHSUDIALI(SU)",
    "lịch sử và địa lí":"LICHSUDIALI","lịch sử và địa lý":"LICHSUDIALI",
    "lich su va dia ly":"LICHSUDIALI","ls&đl":"LICHSUDIALI","ls & đl":"LICHSUDIALI","lsdl":"LICHSUDIALI",
}

# THPT (khối 10–12) — giữ nguyên bảng cũ
SUBJECT_MAP_THPT = {
    "ngữ văn":"NGUVAN","ngữ văn học":"NGUVAN","van":"NGUVAN","nguvan":"NGUVAN","nv":"NGUVAN",
    "toán":"TOAN","toán học":"TOAN","toan":"TOAN",
    "tiếng anh":"ANH","ngoại ngữ 1":"ANH","ngoại ngữ 2":"ANH","ngoại ngữ":"ANH",
    "anh":"ANH","nn1":"ANH","nn2":"ANH","english":"ANH",
    "lịch sử":"LICHSU","lich su":"LICHSU","sử":"LICHSU","su":"LICHSU","lichsu":"LICHSU",
    "giáo dục thể chất":"GDTC","giao duc the chat":"GDTC","thể dục":"GDTC",
    "the duc":"GDTC","gdtc":"GDTC","td":"GDTC",
    "giáo dục quốc phòng và an ninh":"GDQP","giáo dục quốc phòng":"GDQP",
    "giao duc quoc phong":"GDQP","quốc phòng":"GDQP","quoc phong":"GDQP",
    "qpan":"GDQP","gdqp":"GDQP",
    "địa lí":"DIALY","địa lý":"DIALY","dia li":"DIALY","dia ly":"DIALY",
    "địa":"DIALY","dia":"DIALY","dialy":"DIALY",
    "giáo dục kinh tế và pháp luật":"GDKTPL","kinh tế pháp luật":"GDKTPL",
    "kinh te phap luat":"GDKTPL","gdktpl":"GDKTPL","ktpl":"GDKTPL",
    "vật lí":"VATLY","vật lý":"VATLY","vat li":"VATLY","vat ly":"VATLY",
    "lí":"VATLY","lý":"VATLY","li":"VATLY","ly":"VATLY","vatly":"VATLY","vl":"VATLY",
    "hóa học":"HOAHOC","hoá học":"HOAHOC","hoa hoc":"HOAHOC",
    "hóa":"HOAHOC","hoá":"HOAHOC","hoa":"HOAHOC","hoahoc":"HOAHOC","hh":"HOAHOC",
    "sinh học":"SINH","sinh hoc":"SINH","sinh":"SINH",
    "cnnn":"CONGNGHE(NN)","nông nghiệp":"CONGNGHE(NN)","nong nghiep":"CONGNGHE(NN)",
    "công nghệ (nn)":"CONGNGHE(NN)","công nghệ(nn)":"CONGNGHE(NN)","cong nghe nn":"CONGNGHE(NN)",
    "cncn":"CONGNGHE(CN)","công nghiệp":"CONGNGHE(CN)","cong nghiep":"CONGNGHE(CN)",
    "công nghệ (cn)":"CONGNGHE(CN)","công nghệ(cn)":"CONGNGHE(CN)","cong nghe cn":"CONGNGHE(CN)",
    "công nghệ":"CONGNGHE","cong nghe":"CONGNGHE",
    "tin học":"TINHOC","tin hoc":"TINHOC","tin":"TINHOC","tinhoc":"TINHOC",
    "nội dung giáo dục của địa phương":"NDGDDP",
    "nội dung giáo dục địa phương":"NDGDDP","giáo dục địa phương":"NDGDDP",
    "giao duc dia phuong":"NDGDDP","gdđp":"NDGDDP","gddp":"NDGDDP",
    "gd dp":"NDGDDP","gd11dp":"NDGDDP","nd gd dp":"NDGDDP",
    "hoạt động trải nghiệm, hướng nghiệp":"TNHN","hoạt động trải nghiệm":"TNHN",
    "hoat dong trai nghiem":"TNHN","hướng nghiệp":"TNHN","huong nghiep":"TNHN",
    "hđ trải nghiệm":"TNHN","hđtn":"TNHN","hdtn":"TNHN","hđtn hn":"TNHN","tnhn":"TNHN",
    "tiếng pháp":"TIENGPHAP","tieng phap":"TIENGPHAP","pháp":"TIENGPHAP",
    "tiếng nga":"TIENGNGA","tieng nga":"TIENGNGA",
    "tiếng nhật":"TIENGNHAT","tieng nhat":"TIENGNHAT",
    "tiếng trung":"TIENGTRUNG","tieng trung":"TIENGTRUNG",
    "tiếng hàn":"TIENGHAN","tieng han":"TIENGHAN",
    "nghề phổ thông":"NGHEPHOTHONG","nghe pho thong":"NGHEPHOTHONG","nghề":"NGHEPHOTHONG",
    "âm nhạc":"AMNHAC","am nhac":"AMNHAC","nhạc":"AMNHAC","nhac":"AMNHAC",
    "mỹ thuật":"MYTHUAT","mĩ thuật":"MYTHUAT","my thuat":"MYTHUAT",
    "mi thuat":"MYTHUAT","mt":"MYTHUAT",
    "lịch sử và địa lí":"LICHSUDIALI","lịch sử và địa lý":"LICHSUDIALI",
    "lich su va dia ly":"LICHSUDIALI","ls&đl":"LICHSUDIALI",
    "ls & đl":"LICHSUDIALI","lsdl":"LICHSUDIALI",
    "khoa học tự nhiên":"KHTN","khoa hoc tu nhien":"KHTN","khtn":"KHTN",
    "giáo dục công dân":"GDKTPL","giao duc cong dan":"GDKTPL","gdcd":"GDKTPL",
    "hoạt động ngoài giờ lên lớp":"HDNGLL",
    "hoat dong ngoai gio len lop":"HDNGLL","hđngll":"HDNGLL","hdngll":"HDNGLL",
    "tiếng dân tộc thiểu số":"TDTTS","tieng dan toc thieu so":"TDTTS",
    "nghệ thuật":"NGHETHUAT","nghe thuat":"NGHETHUAT",
}

# Alias ngắn để dùng trong code (backward compat)
SUBJECT_MAP = SUBJECT_MAP_THPT

CAP_HOC_OPTIONS = ["THPT", "THCS", "TH"]   # TH = Tiểu học

def _get_subject_map(cap_hoc: str) -> dict:
    """Trả về bảng môn học theo cấp học."""
    if cap_hoc == "TH":   return SUBJECT_MAP_TH
    if cap_hoc == "THCS": return SUBJECT_MAP_THCS
    return SUBJECT_MAP_THPT



def _remove_accent(s):
    s = unicodedata.normalize("NFD", s)
    return "".join(c for c in s if unicodedata.category(c) != "Mn").lower().strip()

# Tổng hợp tất cả mã môn từ cả 3 cấp (dùng cho UI dropdown)
_ALL_CODES = sorted(set(
    list(SUBJECT_MAP_TH.values()) +
    list(SUBJECT_MAP_THCS.values()) +
    list(SUBJECT_MAP_THPT.values())
))

def _build_map_no_accent(smap: dict) -> dict:
    m = {_remove_accent(k): v for k, v in smap.items()}
    for code in set(smap.values()):
        m[code.lower()] = code
    return m

_MAP_NO_ACCENT_TH   = _build_map_no_accent(SUBJECT_MAP_TH)
_MAP_NO_ACCENT_THCS = _build_map_no_accent(SUBJECT_MAP_THCS)
_MAP_NO_ACCENT_THPT = _build_map_no_accent(SUBJECT_MAP_THPT)
# Alias backward compat
_MAP_NO_ACCENT = _MAP_NO_ACCENT_THPT

_fuzzy_cache: dict = {}

def match_subject_local(raw, cap_hoc: str = "THPT"):
    """Tra mã môn học theo cấp học. cap_hoc: 'TH' | 'THCS' | 'THPT'."""
    if not raw: return None
    smap    = _get_subject_map(cap_hoc)
    mna     = _build_map_no_accent(smap)   # nhẹ, dict nhỏ
    s       = raw.lower().strip()
    sn      = _remove_accent(raw)
    if s  in smap:      return smap[s]
    if sn in mna:       return mna[sn]
    if s.upper() in set(smap.values()): return s.upper()
    best, best_len = None, 0
    for key, code in smap.items():
        if key in s or s in key:
            if len(key) > best_len: best, best_len = code, len(key)
    if best: return best
    for key_nd, code in mna.items():
        if key_nd in sn or sn in key_nd:
            if len(key_nd) > best_len: best, best_len = code, len(key_nd)
    if best: return best
    cache_key = (cap_hoc, sn)
    if cache_key in _fuzzy_cache: return _fuzzy_cache[cache_key]
    matches = difflib.get_close_matches(sn, list(mna.keys()), n=1, cutoff=0.72)
    result = mna[matches[0]] if matches else None
    _fuzzy_cache[cache_key] = result
    return result

def get_subject_code(raw, cap_hoc: str = "THPT"):
    return match_subject_local(raw.strip(), cap_hoc) if raw and raw.strip() else None

def _get_cap_hoc_by_grade(grade: int) -> str:
    """Trả về cấp học tương ứng với khối lớp."""
    if grade is None:   return "THPT"
    if grade <= 3:      return "TH"
    if grade <= 5:      return "TH"   # TH khối 4-5 (GVCN subjects khác nhưng cùng bảng TH)
    if grade <= 9:      return "THCS"
    return "THPT"

def get_subject_code_for_class(raw: str, class_name: str) -> str:
    """
    Tra mã môn học dựa trên tên lớp thực tế.
    Khối 1-5 → bảng TH, khối 6-9 → bảng THCS, khối 10-12 → bảng THPT.
    """
    grade = get_grade(class_name) if class_name else None
    cap   = _get_cap_hoc_by_grade(grade)
    return get_subject_code(raw, cap)

_GRADE_PFX  = r'(?:0?[1-9]|1[0-2])'
_CLASS_PAT  = r'(?<!\d)' + _GRADE_PFX + r'[A-Za-zÀ-ỹ]+\.?\d{0,3}(?:\.\d+)?'
_SUBJECT_STOPWORDS = {
    "đến","den","và","va","từ","tu","lớp","lop",
    "khối","khoi","tới","toi","to","the","from"
}

def _enumerate_splits(grade, alpha, digits, known_classes):
    """
    Liệt kê TẤT CẢ các cách tách hợp lệ của chuỗi `digits` sau prefix `grade+alpha`,
    trong đó mỗi phần đều thuộc known_classes.
    Trả về list of list[str], mỗi phần tử là một cách tách.
    Chỉ trả về các cách mà TOÀN BỘ mảnh đều nằm trong known_classes.
    """
    n = len(digits)
    results = []

    def backtrack(pos, current):
        if pos == n:
            results.append(list(current))
            return
        for length in range(1, n - pos + 1):
            candidate = f"{grade}{alpha}{digits[pos:pos+length]}"
            if candidate in known_classes:
                current.append(candidate)
                backtrack(pos + length, current)
                current.pop()

    backtrack(0, [])
    return results


def _is_ambiguous(grade, alpha, digits, known_classes):
    """
    Kiểm tra xem chuỗi có nhiều hơn 1 cách tách hợp lệ không.
    """
    splits = _enumerate_splits(grade, alpha, digits, known_classes)
    return splits if len(splits) > 1 else None


def _split_alpha_compact(grade, alphas, known_classes):
    """
    Tách chuỗi chữ cái liên tiếp thành các lớp riêng lẻ dựa trên known_classes.
    Ví dụ: grade='7', alphas='ABCD', known={'7A','7B','7C','7D'} → ['7A','7B','7C','7D']
    Dùng backtracking (greedy từ trái) để tìm cách tách hợp lệ.
    Trả về list lớp nếu tách được, None nếu không tách được.
    """
    if not known_classes or not alphas:
        return None
    n = len(alphas)
    result = []
    i = 0
    while i < n:
        matched = False
        # Greedy: thử từ dài nhất → ngắn nhất
        for length in range(n - i, 0, -1):
            candidate = f"{grade}{alphas[i:i+length]}"
            if candidate in known_classes:
                result.append(candidate)
                i += length
                matched = True
                break
        if not matched:
            return None   # không thể tách hoàn toàn
    return result if result else None


def expand_class_range(text, known_classes=None, resolved_ambiguities=None):
    """
    Parse chuỗi lớp học thành danh sách lớp.
    Hỗ trợ khối 1-12 (bao gồm dạng viết tắt 1A, 6B, 9C và dạng 01A…).
    Hỗ trợ dạng compact chữ cái: '7ABCD' → ['7A','7B','7C','7D'] khi known_classes có sẵn.
    known_classes: tập hợp tên lớp hợp lệ (từ cột GVCN).
    resolved_ambiguities: dict {raw_token: [lớp đã chọn]} — lựa chọn của người dùng.
    """
    if resolved_ambiguities is None:
        resolved_ambiguities = {}

    # Bỏ sĩ số trong ngoặc: 10A1(52) → 10A1
    text = re.sub(r'((?:0?[1-9]|1[0-2])[A-Za-zÀ-ỹ]+\d*)\(\d+\)', r'\1', text, flags=re.UNICODE)

    # Mở rộng suffix groups (cả số lẫn chữ): 7A,B,C,D → 7A,7B,7C,7D / 11A4,5 → 11A4,11A5
    text = _expand_suffix_groups_in_text(text)

    classes = []

    # ── Xử lý compact chữ cái: 7ABCD → [7A,7B,7C,7D] (khi có known_classes) ─
    # Pattern: khối + 2+ chữ cái + KHÔNG có số → dạng compact alpha
    _alpha_compact_pat = re.compile(
        r'(?<![A-Za-z\d])(0?[1-9]|1[0-2])([A-Za-zÀ-ỹ]{2,})(?!\d)', re.UNICODE)
    def _expand_alpha(m):
        grade, alphas = m.group(1), m.group(2)
        token = f"{grade}{alphas}"
        # Nếu token đã là 1 lớp hợp lệ trong known → giữ nguyên, không tách
        if known_classes and token in known_classes:
            classes.append(token)
            return ''
        if known_classes:
            split_result = _split_alpha_compact(grade, alphas, known_classes)
            if split_result:
                classes.extend(split_result)
                return ''
        # Không tách được → để lại cho regex bình thường bên dưới xử lý
        return m.group()
    text = _alpha_compact_pat.sub(_expand_alpha, text)

    # ── Xử lý range: 1A1-1A5, 9B1 đến 9B3, 10A1-10A5 ──────────────────────
    rp = re.compile(
        r'(0?[1-9]|1[0-2])([A-Za-zÀ-ỹ]+)(\d+)\s*(?:đến|den|-)\s*\1\2(\d+)',
        re.UNICODE)
    for m in rp.finditer(text):
        g, a, s, e = m.groups()
        for i in range(int(s), int(e)+1):
            classes.append(f"{g}{a}{i}")
    text = rp.sub('', text)

    def _split_digits(grade, alpha, digits):
        raw_token = f"{grade}{alpha}{digits}"
        if raw_token in resolved_ambiguities:
            return resolved_ambiguities[raw_token]
        if known_classes:
            splits = _enumerate_splits(grade, alpha, digits, known_classes)
            if len(splits) == 1:
                return splits[0]
            elif len(splits) > 1:
                # Ambiguous → greedy từ trái, UI sẽ hỏi ở lượt sau
                result, i = [], 0
                while i < len(digits):
                    matched = False
                    for length in range(len(digits) - i, 0, -1):
                        candidate = f"{grade}{alpha}{digits[i:i+length]}"
                        if candidate in known_classes:
                            result.append(candidate); i += length; matched = True; break
                    if not matched:
                        result.append(f"{grade}{alpha}{digits[i]}"); i += 1
                return result
            else:
                return [f"{grade}{alpha}{x}" for x in digits]
        else:
            return [f"{grade}{alpha}{x}" for x in digits]

    def _compact(m):
        g, a, d = m.group(1), m.group(2), m.group(3)
        # Leading zero → 1 lớp duy nhất (1A02 = lớp 1A02)
        if d.startswith('0'):
            c = f"{g}{a}{d}"
            if c not in classes: classes.append(c)
            return ''
        # Trailing zero → ưu tiên coi là 1 lớp nếu khớp known hoặc không có known
        if d.endswith('0') and (not known_classes or f"{g}{a}{d}" in known_classes):
            c = f"{g}{a}{d}"
            if c not in classes: classes.append(c)
            return ''
        # 2 chữ số giống nhau (11, 22…) → rõ ràng là 1 lớp, không tách
        # Ví dụ: 11A11 = lớp 11A11, không ai viết 11A11 để chỉ 11A1+11A1
        if len(d) == 2 and d[0] == d[1]:
            c = f"{g}{a}{d}"
            if c not in classes: classes.append(c)
            return ''
        for c in _split_digits(g, a, d):
            if c not in classes: classes.append(c)
        return ''

    # Compact 3+ chữ số: 10A123 → tách, 1A123 → tách (luôn là compact)
    text = re.sub(r'(?<![\w,;])(0?[1-9]|1[0-2])([A-Za-z]+)(\d{3,})', _compact, text)
    # Compact 2 chữ số: chỉ chạy khi có known_classes (để phân biệt 10A12=1 lớp vs 10A1+10A2)
    # Không có known_classes: không thể phân biệt → coi 2 digits là 1 lớp (an toàn hơn)
    if known_classes:
        text = re.sub(r'(?<![\w,;])(0?[1-9]|1[0-2])([A-Za-z]+)(\d{2})(?![,;.\s\d])', _compact, text)

    # ── Xử lý suffix groups: 11A4,5,11,12A6,7 → 11A4,11A5,11A11,12A6,12A7 ──
    # Tokenize thành các nhóm (base, [suffixes]) rồi sinh lớp
    _SUFFIX_TOK = re.compile(
        r'(?P<cls_full>(?:0?[1-9]|1[0-2])[A-Za-zÀ-ỹ]+\d+)'   # lớp đầy đủ: 11A4
        r'|(?P<cls_pfx>(?:0?[1-9]|1[0-2])[A-Za-zÀ-ỹ]+)'       # prefix: 11A
        r'|(?P<num>\d+)'                                          # số suffix: 5, 11
        r'|(?P<sep>[,;\s]+)'                                      # dấu phân cách
        r'|(?P<other>.)',
        re.UNICODE
    )
    def _parse_suffix_groups(seg):
        """
        Từ chuỗi '11A4,5,11,12A6,7' trả về [('11A',['4','5','11']),('12A',['6','7'])].
        cls_full bắt đầu nhóm mới (base = phần chữ, digit đầu = suffix[0]).
        num tiếp theo là suffix cho base hiện tại.
        """
        groups, cur_base, cur_nums = [], None, []
        for kind, val in ((m.lastgroup, m.group().strip())
                          for m in _SUFFIX_TOK.finditer(seg)):
            if not val or kind in ('sep', 'other'): continue
            if kind == 'cls_full':
                if cur_base and cur_nums: groups.append((cur_base, cur_nums))
                bm = re.match(r'((?:0?[1-9]|1[0-2])[A-Za-zÀ-ỹ]+)(\d+)$', val, re.UNICODE)
                cur_base, cur_nums = (bm.group(1), [bm.group(2)]) if bm else (val, [])
            elif kind == 'cls_pfx':
                if cur_base and cur_nums: groups.append((cur_base, cur_nums))
                cur_base, cur_nums = val, []
            elif kind == 'num' and cur_base is not None:
                cur_nums.append(val)
        if cur_base and cur_nums: groups.append((cur_base, cur_nums))
        return groups

    # ── Apply suffix groups: 11A4,5,11,12A6,7 → 11A4,11A5,11A11,12A6,12A7 ──
    # Bắt cụm: lớp đầy đủ + (phẩy + số-hoặc-lớp) lặp lại ≥1 lần
    # CLASS alternation đứng trước \d+ để tránh nhầm grade prefix là suffix digit
    _SFX_SEG = re.compile(
        r'(?:0?[1-9]|1[0-2])[A-Za-zÀ-ỹ]+\d+'
        r'(?:\s*[,;]\s*'
        r'(?:(?:0?[1-9]|1[0-2])[A-Za-zÀ-ỹ]+\d+|\d+)'
        r')+',
        re.UNICODE
    )
    def _apply_sfx(m):
        for base, nums in _parse_suffix_groups(m.group()):
            for n in nums:
                c = f"{base}{n}"
                if c not in classes: classes.append(c)
        return ''
    text = _SFX_SEG.sub(_apply_sfx, text)

    classes.extend(re.findall(_CLASS_PAT, text))
    result, seen = [], set()
    for c in classes:
        c = c.strip().strip(',').strip()
        if c and c not in seen: seen.add(c); result.append(c)
    return result


def detect_ambiguous_in_data(df, col_pccm, col_gvcn, known_classes):
    """
    Quét toàn bộ dữ liệu PCCM, tìm tất cả chuỗi token ambiguous.
    Trả về list of dict:
      {
        "token":       "10A123",          # chuỗi gốc trong file
        "grade":       "10",
        "alpha":       "A",
        "digits":      "123",
        "splits":      [["10A1","10A2","10A3"], ["10A12","10A3"], ...],
        "occurrences": ["GV Nguyễn Văn A (Văn: 10A123)", ...]   # mô tả ngữ cảnh
      }
    Mỗi token duy nhất chỉ xuất hiện 1 lần trong kết quả.
    """
    if not known_classes:
        return []

    _ambig_pat = re.compile(r'(0?[1-9]|1[0-2])([A-Za-z]+)(\d{2,})', re.UNICODE)
    found = {}   # token → dict

    col_hoten = None
    for cand in ["họ tên","họ và tên","tên","giáo viên","ho ten","hoten"]:
        col_hoten = find_column(df, [cand])
        if col_hoten: break

    for _, row in df.iterrows():
        praw = str(row.get(col_pccm, "")).strip() if pd.notna(row.get(col_pccm)) else ""
        if not praw:
            continue
        hoten = str(row.get(col_hoten, "")).strip() if col_hoten else ""

        # Chuẩn hoá sơ bộ giống parse_pccm
        text = re.sub(r'\([^)]*\)', '', praw)
        text = text.replace(';', ',').replace('\n', '+')
        # Xoá range đã rõ (1A1-1A5, 10A1-10A5, 9B1 đến 9B4)
        text = re.sub(r'(?:0?[1-9]|1[0-2])[A-Za-z]+\d+\s*(?:đến|den|-)\s*(?:0?[1-9]|1[0-2])[A-Za-z]+\d+', '', text)

        for m in _ambig_pat.finditer(text):
            grade, alpha, digits = m.group(1), m.group(2), m.group(3)
            # Chỉ xét nếu chuỗi ≥ 2 chữ số (10A12 trở lên)
            if len(digits) < 2:
                continue
            token = f"{grade}{alpha}{digits}"
            # Nếu token CHÍNH LÀ 1 lớp hợp lệ trong known → không ambiguous
            if token in known_classes:
                continue
            splits = _enumerate_splits(grade, alpha, digits, known_classes)
            if len(splits) <= 1:
                continue  # không ambiguous

            # Tìm ngữ cảnh môn học xung quanh token
            context_start = max(0, m.start() - 20)
            context = "..." + text[context_start:m.end()+5].strip() + "..."
            occurrence = f"{hoten}: …{context}…" if hoten else context

            if token not in found:
                found[token] = {
                    "token": token,
                    "grade": grade,
                    "alpha": alpha,
                    "digits": digits,
                    "splits": splits,
                    "occurrences": [occurrence],
                }
            else:
                if occurrence not in found[token]["occurrences"]:
                    found[token]["occurrences"].append(occurrence)

    return list(found.values())


def detect_unknown_subjects(df, col_pccm, cap_hoc: str = "AUTO"):
    """
    Quét cột PCCM, tìm tất cả tên môn không nhận diện được.
    Khi cap_hoc="AUTO" (mặc định mới), tra theo bảng môn của từng khối lớp.
    Trả về list of dict: {"raw", "suggestion", "occurrences"}.
    """
    col_hoten = None
    for cand in ["họ tên","họ và tên","tên","giáo viên","ho ten","hoten"]:
        col_hoten = find_column(df, [cand])
        if col_hoten: break

    found = {}  # raw_lower → dict

    for _, row in df.iterrows():
        praw = str(row.get(col_pccm, "")).strip() if pd.notna(row.get(col_pccm)) else ""
        if not praw:
            continue
        hoten = str(row.get(col_hoten, "")).strip() if col_hoten else ""

        parsed = parse_pccm(praw)
        for sr, ll in parsed:
            if not sr:
                continue
            # Tra môn: dùng khối của lớp đầu tiên nếu AUTO, nếu không dùng cap_hoc cố định
            if cap_hoc == "AUTO":
                first_cls = ll[0].strip() if ll else ""
                grade     = get_grade(first_cls) if first_cls else None
                level     = _get_cap_hoc_by_grade(grade)
            else:
                level = cap_hoc
            code = get_subject_code(sr, level)
            if code:
                continue  # đã nhận diện được
            key = sr.lower().strip()
            # Gợi ý: thử tất cả 3 bảng
            suggestion = None
            sn = _remove_accent(sr)
            for lvl in ("TH", "THCS", "THPT"):
                mna = _build_map_no_accent(_get_subject_map(lvl))
                matches = difflib.get_close_matches(sn, list(mna.keys()), n=1, cutoff=0.50)
                if matches:
                    suggestion = mna[matches[0]]
                    break
            ctx = (f"{hoten}: {sr} → {', '.join(ll[:3])}{'…' if len(ll)>3 else ''}"
                   if hoten else f"{sr} → {', '.join(ll[:3])}")
            if key not in found:
                found[key] = {"raw": sr, "suggestion": suggestion, "occurrences": [ctx]}
            else:
                if ctx not in found[key]["occurrences"]:
                    found[key]["occurrences"].append(ctx)

    return list(found.values())

def _expand_suffix_groups_in_text(text):
    """
    Tiền xử lý: mở rộng suffix groups ngay trong chuỗi text TRƯỚC KHI tokenize.
    Xử lý cả suffix số lẫn suffix chữ cái:
      '11A4,5,11,12A6,7' → '11A4,11A5,11A11,12A6,12A7'   (số suffix)
      '10A1,2,3'         → '10A1,10A2,10A3'               (số suffix)
      '7A,B,C,D'         → '7A,7B,7C,7D'                  (chữ suffix - MỚI)
      '11A3, 12D'        → '11A3, 12D'  (12D = lớp riêng)
    """
    # ── Bước 0: mở rộng alpha suffix: 7A,B,C,D → 7A,7B,7C,7D ───────────────
    # Phải chạy trước bước số để tránh tokenizer bị nhầm
    _GP0 = r'(?:0?[1-9]|1[0-2])'
    def _expand_alpha_suffix(m):
        base_cls = m.group(1)   # vd: '7A'
        grade    = re.match(r'(' + _GP0 + r')', base_cls).group(1)
        parts    = [base_cls]
        for ch in re.findall(r'[A-Za-zÀ-ỹ]', m.group(2)):
            parts.append(f"{grade}{ch}")
        return ','.join(parts)
    # Pattern: lớp không số + (phẩy + 1 chữ cái đơn) lặp lại ≥1
    text = re.sub(
        r'(' + _GP0 + r'[A-Za-zÀ-ỹ]+(?!\d))'          # lớp không số: 7A, 7AB...
        r'((?:\s*[,;]\s*[A-Za-zÀ-ỹ](?![A-Za-zÀ-ỹ\d]))+)',  # ,B ,C ,D ...
        _expand_alpha_suffix,
        text,
        flags=re.UNICODE
    )
    _GP = r'(?:0?[1-9]|1[0-2])'
    # Tokenizer phân biệt 3 loại: lớp có số cuối, lớp không có số, số thuần
    _TOK = re.compile(
        r'(?P<cls_full>(?:' + _GP + r')[A-Za-zÀ-ỹ]+\d+)'       # lớp có số: 11A3, 12A6
        r'|(?P<cls_nodig>(?:' + _GP + r')[A-Za-zÀ-ỹ]+(?!\d))'   # lớp không số: 12D, 11D
        r'|(?P<num>\d+)'                                           # số suffix: 5, 11
        r'|(?P<sep>[,;\s]+)'
        r'|(?P<oth>.)',
        re.UNICODE
    )
    tokens = [(m.lastgroup, m.group().strip(), m.start(), m.end())
              for m in _TOK.finditer(text)]

    runs = []   # (start, end, expanded_str) cho các đoạn có suffix thực sự
    j = 0
    while j < len(tokens):
        kind, val, ts, te = tokens[j]
        if kind != 'cls_full':
            j += 1; continue
        bm = re.match(r'((?:' + _GP + r')[A-Za-zÀ-ỹ]+)(\d+)$', val, re.UNICODE)
        if not bm:
            j += 1; continue

        base, parts = bm.group(1), [bm.group(2)]
        run_start, run_end = ts, te
        k = j + 1

        while k < len(tokens):
            k2, v2, ts2, te2 = tokens[k]
            if k2 == 'sep':
                k += 1; continue
            if k2 == 'num':
                parts.append(v2); run_end = te2; k += 1
            elif k2 == 'cls_full':
                bm2 = re.match(r'((?:' + _GP + r')[A-Za-zÀ-ỹ]+)(\d+)$', v2, re.UNICODE)
                if bm2 and bm2.group(1) != base:
                    # Lớp mới khác base → flush nhóm cũ, bắt đầu nhóm mới
                    if len(parts) > 1:
                        runs.append((run_start, run_end,
                                     ','.join(f'{base}{p}' for p in parts)))
                    base, parts = bm2.group(1), [bm2.group(2)]
                    run_start, run_end = ts2, te2
                elif bm2:
                    parts.append(bm2.group(2)); run_end = te2
                k += 1
            elif k2 == 'cls_nodig':
                break   # lớp không số (12D, 11D) → kết thúc run
            else:
                break

        if len(parts) > 1:
            runs.append((run_start, run_end,
                         ','.join(f'{base}{p}' for p in parts)))
        j = k if k > j else j + 1

    if not runs:
        return text
    result, pos = [], 0
    for start, end, expanded in runs:
        result.append(text[pos:start])
        result.append(expanded)
        pos = end
    result.append(text[pos:])
    return ''.join(result)

def parse_pccm(raw_pccm, known_classes=None, resolved_ambiguities=None):
    if not raw_pccm or (isinstance(raw_pccm,float) and pd.isna(raw_pccm)): return []
    text = str(raw_pccm).strip()
    def ep(m):
        inner=m.group(1).strip()
        return '' if re.fullmatch(r'\d+',inner) else ','+inner+','
    text = re.sub(r'\(([^)]*)\)', ep, text)
    text = text.replace(';',',').replace('\n','+')
    # ── Tiền xử lý suffix groups trước khi tokenize ──────────────────────────
    # Ví dụ: '11A4,5,11,12A6,7' → '11A4,11A5,11A11,12A6,12A7'
    text = _expand_suffix_groups_in_text(text)
    # CRP: class-range pattern — hỗ trợ khối 1-12
    _GP = r'(?:0?[1-9]|1[0-2])'   # grade prefix (local)
    CRP = (r''+_GP+r'[A-Za-zÀ-ỹ]+\d+\s*(?:đến|den|-)\s*'+_GP+r'[A-Za-zÀ-ỹ]+\d+'   # range
           r'|(?<!\d)'+_GP+r'[A-Za-zÀ-ỹ]+\d{3,}'                                    # compact 3+
           r'|'+_CLASS_PAT)                                                            # normal
    tokens,results = [],[]
    tr = re.compile(r'(?P<class>'+CRP+r')|(?P<sep>[+,\s]+)|(?P<colon>:)'
                    r'|(?P<word>[A-Za-zÀ-ỹĐđ][A-Za-zÀ-ỹĐđ\(\)]*)|(?P<other>.)',re.UNICODE)
    for m in tr.finditer(text): tokens.append((m.lastgroup, m.group().strip()))
    merged,i = [],0
    while i < len(tokens):
        kind,val = tokens[i]
        if kind=='word':
            words=[val]; j=i+1
            while j<len(tokens):
                k2,v2=tokens[j]
                if k2=='word': words.append(v2); j+=1
                elif k2=='sep' and j+1<len(tokens) and tokens[j+1][0]=='word':
                    words.append(tokens[j+1][1]); j+=2
                # Gom số đứng sau word (ví dụ: "Ngoại ngữ" + "1" → "Ngoại ngữ 1")
                elif k2=='other' and re.fullmatch(r'\d+', v2):
                    words.append(v2); j+=1
                elif k2=='sep' and j+1<len(tokens) and tokens[j+1][0]=='other' and re.fullmatch(r'\d+', tokens[j+1][1]):
                    words.append(tokens[j+1][1]); j+=2
                else: break
            merged.append(('word',' '.join(words))); i=j
        elif kind=='sep':
            if val: merged.append(('sep',val))
            i+=1
        elif kind in ('class','colon','other'): merged.append((kind,val)); i+=1
        else: i+=1
    cur_subj,cur_cls = None,[]
    def flush(s,c,o):
        if s and c: o.append((s,c))
        elif c: o.append(("",c))
    idx=0
    while idx<len(merged):
        kind,val = merged[idx]
        if kind=='word':
            if _remove_accent(val) in _SUBJECT_STOPWORDS: idx+=1; continue
            nns=None
            for k2,v2 in merged[idx+1:]:
                if k2!='sep': nns=(k2,v2); break
            if nns and nns[0]=='colon':
                flush(cur_subj,cur_cls,results); cur_subj=val; cur_cls=[]
                idx+=1
                while idx<len(merged) and merged[idx][0] in ('sep','colon'): idx+=1
            elif nns and nns[0]=='class':
                flush(cur_subj,cur_cls,results); cur_subj=val; cur_cls=[]
                idx+=1
                while idx<len(merged) and merged[idx][0]=='sep': idx+=1
            else: idx+=1
        elif kind=='class':
            cur_cls.extend(expand_class_range(val, known_classes, resolved_ambiguities))
            idx+=1
        elif kind in ('sep','colon','other'): idx+=1
        else: idx+=1
    flush(cur_subj,cur_cls,results)
    return results

def format_date(val):
    try:
        if val is None: return None,""
        if isinstance(val,datetime): return val,val.strftime("%d/%m/%Y")
        if isinstance(val,date_type):
            dt=datetime(val.year,val.month,val.day); return dt,dt.strftime("%d/%m/%Y")
        if isinstance(val,(int,float)):
            if pd.isna(val): return None,""
            dt=datetime(1899,12,30)+timedelta(days=int(val)); return dt,dt.strftime("%d/%m/%Y")
        s=str(val).strip()
        if not s or s.lower() in ("nan","nat","none",""): return None,""
        for fmt in ("%d/%m/%Y","%Y-%m-%d","%d-%m-%Y","%m/%d/%Y","%d/%m/%y","%Y/%m/%d"):
            try: dt=datetime.strptime(s,fmt); return dt,dt.strftime("%d/%m/%Y")
            except: pass
        return None,s
    except: return None,""

def find_column(df,candidates):
    cl={c.lower().strip():c for c in df.columns}
    for cand in candidates:
        c=cand.lower().strip()
        if c in cl: return cl[c]
        for key,orig in cl.items():
            if c in key or key in c: return orig
    return None

def detect_header_row(sdf):
    kws=['stt','họ tên','họ và tên','giáo viên','pccm','phân công','ngày sinh']
    for i,row in sdf.iterrows():
        vals=[str(v).lower().strip() for v in row.values if pd.notna(v)]
        if sum(1 for v in vals for k in kws if k in v)>=2: return i
    return 0

def build_known_classes_from_gvcn(df, col_gvcn) -> set:
    """
    Thu thập danh sách lớp chuẩn từ cột GVCN, xử lý mọi dạng viết tắt:
      - '7A'        → {7A}
      - '7ABCD'     → {7A,7B,7C,7D}   compact alpha liên tiếp
      - '7A,B,C,D'  → {7A,7B,7C,7D}   alpha suffix sau dấu phẩy
      - '10A1'      → {10A1}           lớp có số, giữ nguyên
      - '11A12'     → {11A12}          giữ nguyên, không tách
    Trả về set rỗng nếu col_gvcn không tồn tại hoặc toàn bộ ô trống.
    """
    known: set = set()
    if not col_gvcn:
        return known
    _GP = r'(?:0?[1-9]|1[0-2])'
    _TOK = re.compile(
        r'(?P<full>'   + _GP + r'[A-Za-zÀ-ỹ]+\d+)'            # lớp đầy đủ: 10A1, 11A12
        r'|(?P<base>'  + _GP + r'[A-Za-zÀ-ỹ]{2,}(?!\d))'      # compact alpha: 7ABCD, 7AB
        r'|(?P<single>'+ _GP + r'[A-Za-zÀ-ỹ](?![A-Za-zÀ-ỹ\d]))'  # khối+1 chữ: 7A
        r'|(?P<alpha>[A-Za-zÀ-ỹ](?![A-Za-zÀ-ỹ\d]))'           # alpha suffix: B, C
        r'|(?P<sep>[,;\s]+)'
        r'|(?P<other>.)',
        re.UNICODE
    )
    for raw_val in df[col_gvcn]:
        val = str(raw_val).strip() if pd.notna(raw_val) else ''
        if not val or val.lower() in ('nan', 'none', ''):
            continue
        cur_base = None
        for m in _TOK.finditer(val):
            kind, v = m.lastgroup, m.group().strip()
            if not v or kind == 'sep':
                continue
            elif kind == 'full':
                known.add(v)
                cur_base = None
            elif kind == 'base':
                grade = re.match(r'(0?[1-9]|1[0-2])', v).group(1)
                for ch in v[len(grade):]:
                    known.add(f"{grade}{ch}")
                cur_base = grade
            elif kind == 'single':
                known.add(v)
                cur_base = re.match(r'(0?[1-9]|1[0-2])', v).group(1)
            elif kind == 'alpha':
                if cur_base:
                    known.add(f"{cur_base}{v}")
            else:
                cur_base = None
    return known


def get_grade(cls):
    """Trả về số khối (int) từ tên lớp. Hỗ trợ khối 1-12 và dạng 01A, 09B."""
    m = re.match(r'^(0?[1-9]|1[0-2])(?=[A-Za-zÀ-ỹ])', str(cls).strip(), re.UNICODE)
    return int(m.group(1)) if m else None

def _sh(ws,row,ncols,color="1F4E79"):
    fill=PatternFill("solid",fgColor=color)
    font=Font(bold=True,color="FFFFFF",name="Arial",size=11)
    align=Alignment(horizontal="center",vertical="center",wrap_text=True)
    for col in range(1,ncols+1):
        cell=ws.cell(row=row,column=col)
        cell.fill=fill; cell.font=font; cell.alignment=align

def _sdr(ws,row,ncols,even,left_cols=()):
    fill=PatternFill("solid",fgColor="EBF3FB" if even else "FFFFFF")
    font=Font(name="Arial",size=10)
    for col in range(1,ncols+1):
        cell=ws.cell(row=row,column=col)
        cell.fill=fill; cell.font=font
        cell.alignment=(Alignment(horizontal="left",vertical="center",wrap_text=True)
                        if col in left_cols
                        else Alignment(horizontal="center",vertical="center"))

def _ab(ws,sr,er,ncols):
    thin=Side(style='thin',color='B0C4DE')
    border=Border(left=thin,right=thin,top=thin,bottom=thin)
    for row in range(sr,er+1):
        for col in range(1,ncols+1): ws.cell(row=row,column=col).border=border

def process_data(input_src, nien_khoa: str, cap_hoc: str = "AUTO",
                 progress_cb=None,
                 resolved_ambiguities=None, resolved_subjects=None) -> bytes:
    """
    Xử lý file Excel đầu vào. KHÔNG cần API key.
    cap_hoc: "AUTO" (mặc định) → tự động tra môn theo khối lớp thực tế.
             "TH" / "THCS" / "THPT" → ép buộc một cấp (backward compat).
    """
    def log(m):
        if progress_cb: progress_cb(m)

    def _resolve_code(sr: str, lop: str) -> str | None:
        """Tra mã môn theo lớp thực tế (AUTO) hoặc cap_hoc cố định."""
        rs_key = sr.lower().strip() if sr else ""
        if resolved_subjects and rs_key in resolved_subjects:
            return resolved_subjects[rs_key]
        if cap_hoc == "AUTO":
            return get_subject_code_for_class(sr, lop)
        return get_subject_code(sr, cap_hoc)

    src = (io.BytesIO(input_src) if isinstance(input_src,(bytes,bytearray))
           else input_src)

    xl = pd.ExcelFile(src)
    ds = next((s for s in xl.sheet_names if s.strip().lower()=="data"), xl.sheet_names[0])
    log(f"Đọc sheet '{ds}'...")
    rdf = pd.read_excel(src,sheet_name=ds,header=None)
    hri = detect_header_row(rdf)
    df  = pd.read_excel(src,sheet_name=ds,header=hri)
    df.columns = [str(c).strip() for c in df.columns]

    col_stt   = find_column(df,["stt","tt","số thứ tự","no"])
    col_hoten = find_column(df,["họ tên","họ và tên","tên","giáo viên","ho ten","hoten"])
    col_ngay  = find_column(df,["ngày sinh","ngay sinh","sinh ngày","dob","birthday"])
    col_pccm  = find_column(df,["pccm","phân công chuyên môn","phân công",
                                 "giảng dạy lớp","môn học giảng dạy","phan cong","giang day"])
    col_gvcn  = find_column(df,["gvcn","chủ nhiệm","chu nhiem","chủ nhiệm lớp",
                                 "chu nhiem lop","lớp chủ nhiệm","lop chu nhiem","cn"])
    if not col_hoten: raise ValueError("Không tìm thấy cột Họ tên!")
    if not col_pccm:  raise ValueError("Không tìm thấy cột PCCM!")

    df = df[df[col_hoten].notna()&(df[col_hoten].astype(str).str.strip()!="")].copy()
    df = df.reset_index(drop=True)

    # ── Bước 1: Thu thập known_classes từ cột GVCN ───────────────────────────
    # build_known_classes_from_gvcn xử lý đầy đủ tất cả các dạng:
    #   7A, 7ABCD → {7A,7B,7C,7D},  7A,B,C,D → {7A,7B,7C,7D},  10A1 → {10A1}
    # Khi cột GVCN trống/không tồn tại → known_classes rỗng,
    # parse_pccm tiếp tục theo logic mặc định.
    known_classes: set = set()
    if col_gvcn:
        log("Đọc danh sách lớp từ cột GVCN...")
        known_classes = build_known_classes_from_gvcn(df, col_gvcn)
        if known_classes:
            log(f"  → {len(known_classes)} lớp: {', '.join(sorted(known_classes))}")
        else:
            log("  → Cột GVCN không có dữ liệu lớp, xử lý PCCM theo logic mặc định.")

    total = len(df)
    teachers = []

    for idx,row in df.iterrows():
        log(f"Xử lý giáo viên {idx+1}/{total}: {row[col_hoten]}")
        stt   = str(row[col_stt]).strip() if col_stt and pd.notna(row.get(col_stt)) else str(idx+1)
        hoten = str(row[col_hoten]).strip()
        ndt,nstr = (format_date(row[col_ngay]) if col_ngay and pd.notna(row.get(col_ngay))
                    else (None,""))
        praw = str(row[col_pccm]).strip() if pd.notna(row.get(col_pccm)) else ""

        # Đọc lớp chủ nhiệm
        # Ưu tiên dùng known_classes (đã xây dựng từ toàn bộ cột GVCN) để tách
        # dạng compact chữ cái: "7ABCD" → [7A, 7B, 7C, 7D]
        cn_classes = []
        gvcn_str   = ""
        if col_gvcn and pd.notna(row.get(col_gvcn)):
            gvcn_raw = str(row[col_gvcn]).strip()
            if gvcn_raw:
                # Dùng known_classes để expand chính xác (bao gồm dạng 7ABCD)
                cn_classes = expand_class_range(
                    gvcn_raw,
                    known_classes if known_classes else None,
                    resolved_ambiguities or {}
                )
                gvcn_str = ", ".join(cn_classes) if cn_classes else gvcn_raw

        # ── Xử lý PCCM ────────────────────────────────────────────────────────
        # Tiểu học AUTO/TH: GVCN tự động dạy các môn chính khóa của lớp chủ nhiệm.
        # Khối được xác định từ lớp CN thực tế (không phụ thuộc cap_hoc truyền vào).
        mllist: list = []

        cn_grade = get_grade(cn_classes[0]) if cn_classes else None
        is_th_gvcn = (cn_grade is not None and 1 <= cn_grade <= 5 and
                      (cap_hoc == "AUTO" or cap_hoc == "TH"))

        if is_th_gvcn:
            cn_subjects = _TH_GVCN_SUBJECTS.get(cn_grade, [])
            if cn_subjects:
                log(f"  → GVCN TH khối {cn_grade}: tự động thêm {cn_subjects}")
            for lop in cn_classes:
                for code in cn_subjects:
                    mllist.append((lop.strip(), code))

        # Xử lý PCCM bình thường — tra môn theo khối lớp thực tế
        parsed = parse_pccm(praw, known_classes if known_classes else None,
                             resolved_ambiguities or {})
        for sr, ll in parsed:
            for lop in ll:
                lop = lop.strip()
                if not lop: continue
                code = _resolve_code(sr, lop)
                if code:
                    mllist.append((lop, code))
                else:
                    mllist.append((lop, sr.upper() if sr else "?"))

        # Dedup giữ thứ tự (CN subjects trước, PCCM sau)
        seen = set(); uml = []
        for lop, code in mllist:
            if (lop, code) not in seen:
                seen.add((lop, code)); uml.append((lop, code))

        # Tổng hợp mã môn
        scodes = []
        for _, code in uml:
            if code and code != "?" and code not in scodes:
                scodes.append(code)

        teachers.append({"stt":stt,"ho_ten":hoten,"ngay_dt":ndt,"ngay_str":nstr,
                         "subject_codes":scodes,"mon_lop_list":uml,"gvcn_str":gvcn_str})

    pc = defaultdict(list)
    for t in teachers:
        for lop,code in t["mon_lop_list"]: pc[(lop,code)].append(t["ho_ten"])

    for t in teachers:
        parts=[]
        for lop,code in t["mon_lop_list"]:
            key=(lop,code)
            parts.append(f"{lop}-{code}({t['ho_ten']})" if len(pc[key])>1 else f"{lop}-{code}")
        t["pccm_str"]=",".join(parts)

    # Thu thập tất cả lớp duy nhất từ cả PCCM lẫn CN
    all_cls_set = set()
    for t in teachers:
        for lop, _ in t["mon_lop_list"]:
            lop = lop.strip()
            if lop: all_cls_set.add(lop)
        if t.get("gvcn_str"):
            for lop in t["gvcn_str"].split(","):
                lop = lop.strip()
                if lop: all_cls_set.add(lop)
    all_cls = sorted(all_cls_set, key=lambda c: (get_grade(c) or 99, c))

    log("Tạo file Excel đầu ra...")
    wb = openpyxl.Workbook()

    # ── Class sheet ───────────────────────────────────────────────────
    wc=wb.active; wc.title="Class"
    wc["A1"]="Niên khóa"; wc["B1"]=nien_khoa
    # wc["A2"]="Cấp học";   wc["B2"]={"TH":"Tiểu học","THCS":"THCS","THPT":"THPT"}.get(cap_hoc, cap_hoc)
    wc["A2"]="Lớp";       wc["B2"]="Khối"
    for r in (1,2):
        for col in ("A","B"):
            c=wc[f"{col}{r}"]
            c.fill=PatternFill("solid",fgColor="1F4E79")
            c.font=Font(bold=True,color="FFFFFF",name="Arial",size=11)
            c.alignment=Alignment(horizontal="center",vertical="center")
    for i,cls in enumerate(all_cls):
        r=i+3
        wc.cell(row=r,column=1,value=cls); wc.cell(row=r,column=2,value=get_grade(cls))
        for col in (1,2):
            c=wc.cell(row=r,column=col)
            c.fill=PatternFill("solid",fgColor="EBF3FB" if i%2==0 else "FFFFFF")
            c.font=Font(name="Arial",size=10)
            c.alignment=Alignment(horizontal="center",vertical="center")
    _ab(wc,1,len(all_cls)+3,2)
    wc.column_dimensions["A"].width=14; wc.column_dimensions["B"].width=14
    wc.freeze_panes="A3"

    # ── Teachers sheet ────────────────────────────────────────────────
    wt=wb.create_sheet("Teachers")
    ht=["STT","Họ tên","Ngày sinh","SĐT","Môn dạy","TBM","CN","PCCM","Email","Tên đăng nhập"]
    for ci,h in enumerate(ht,1): wt.cell(row=1,column=ci,value=h)
    _sh(wt,1,len(ht)); wt.row_dimensions[1].height=30
    for i,t in enumerate(teachers):
        rn=i+2
        wt.cell(row=rn,column=1,value=t["stt"])
        wt.cell(row=rn,column=2,value=t["ho_ten"])
        dc=wt.cell(row=rn,column=3)
        if t["ngay_dt"]: dc.value=t["ngay_dt"]; dc.number_format="DD/MM/YYYY"
        else: dc.value=t["ngay_str"]
        wt.cell(row=rn,column=4,value="")
        wt.cell(row=rn,column=5,value=", ".join(t["subject_codes"]))
        wt.cell(row=rn,column=6,value="")
        wt.cell(row=rn,column=7,value=t.get("gvcn_str",""))
        wt.cell(row=rn,column=8,value=t["pccm_str"])
        _sdr(wt,rn,len(ht),i%2==0,left_cols=(2,5,8))
    _ab(wt,1,len(teachers)+1,len(ht))
    for ci,w in enumerate([6,25,14,14,30,10,10,80],1):
        wt.column_dimensions[get_column_letter(ci)].width=w
    wt.freeze_panes="A2"

    # ── Students sheet ────────────────────────────────────────────────
    ws=wb.create_sheet("Students")
    hs=["STT","Mã HS","Họ tên","Lớp","Giới tính","Ngày sinh","Số điện thoại","Email","Tài khoản"]
    for ci,h in enumerate(hs,1): ws.cell(row=1,column=ci,value=h)
    _sh(ws,1,len(hs)); ws.row_dimensions[1].height=30
    _ab(ws,1,1,len(hs))
    for ci,w in enumerate([6,14,25,10,12,14,16,28,18],1):
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.freeze_panes="A2"

    out=io.BytesIO(); wb.save(out); out.seek(0)
    log("Hoàn thành!")
    return out.read()
